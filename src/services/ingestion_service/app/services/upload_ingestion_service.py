from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any, Literal

from fastapi import Depends, HTTPException, status
from openpyxl import load_workbook
from pydantic import BaseModel, ValidationError

from app.DTOs.business_date_dto import BusinessDate
from app.DTOs.fx_rate_dto import FxRate
from app.DTOs.instrument_dto import Instrument
from app.DTOs.market_price_dto import MarketPrice
from app.DTOs.portfolio_dto import Portfolio
from app.DTOs.transaction_dto import Transaction
from app.DTOs.upload_dto import (
    UploadCommitResponse,
    UploadEntityType,
    UploadPreviewResponse,
    UploadRowError,
)
from app.services.ingestion_service import IngestionService, get_ingestion_service


MODEL_BY_ENTITY: dict[UploadEntityType, type[BaseModel]] = {
    "portfolios": Portfolio,
    "instruments": Instrument,
    "transactions": Transaction,
    "market_prices": MarketPrice,
    "fx_rates": FxRate,
    "business_dates": BusinessDate,
}


def _normalized_key(value: str) -> str:
    return "".join(ch for ch in value.lower() if ch.isalnum())


def _field_alias_index(model_cls: type[BaseModel]) -> dict[str, str]:
    index: dict[str, str] = {}
    for field_name, field_info in model_cls.model_fields.items():
        index[_normalized_key(field_name)] = field_name
        alias = field_info.alias
        if alias:
            index[_normalized_key(alias)] = alias
    return index


def _normalize_row(row: dict[str, Any], alias_index: dict[str, str]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for raw_key, raw_value in row.items():
        if raw_key is None:
            continue
        canonical_key = alias_index.get(_normalized_key(str(raw_key)))
        if canonical_key is None:
            continue
        value = raw_value
        if isinstance(value, str):
            stripped = value.strip()
            value = None if stripped == "" else stripped
        normalized[canonical_key] = value
    return normalized


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records: list[dict[str, Any]] = []
    for row_values in rows[1:]:
        if row_values is None:
            continue
        row_dict = {
            headers[index]: row_values[index] if index < len(row_values) else None
            for index in range(len(headers))
            if headers[index]
        }
        if any(value is not None and str(value).strip() != "" for value in row_dict.values()):
            records.append(row_dict)
    return records


def _detect_format(filename: str) -> Literal["csv", "xlsx"]:
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        return "csv"
    if lowered.endswith(".xlsx"):
        return "xlsx"
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Unsupported file format. Use .csv or .xlsx.",
    )


@dataclass
class _ValidationResult:
    file_format: Literal["csv", "xlsx"]
    valid_models: list[BaseModel]
    valid_rows: list[dict[str, Any]]
    errors: list[UploadRowError]
    total_rows: int


class UploadIngestionService:
    def __init__(self, ingestion_service: IngestionService):
        self._ingestion_service = ingestion_service

    def _parse_rows(
        self, filename: str, content: bytes
    ) -> tuple[Literal["csv", "xlsx"], list[dict[str, Any]]]:
        file_format = _detect_format(filename)
        if file_format == "csv":
            return file_format, _parse_csv(content)
        return file_format, _parse_xlsx(content)

    def _validate_rows(
        self, entity_type: UploadEntityType, filename: str, content: bytes
    ) -> _ValidationResult:
        model_cls = MODEL_BY_ENTITY[entity_type]
        alias_index = _field_alias_index(model_cls)
        file_format, rows = self._parse_rows(filename, content)

        valid_models: list[BaseModel] = []
        valid_rows: list[dict[str, Any]] = []
        errors: list[UploadRowError] = []

        for index, row in enumerate(rows, start=2):
            normalized_row = _normalize_row(row, alias_index)
            try:
                model = model_cls.model_validate(normalized_row)
            except ValidationError as exc:
                issues: list[str] = []
                for error in exc.errors():
                    location = ".".join(str(part) for part in error.get("loc", ()))
                    issues.append(f"{location}: {error.get('msg', 'invalid value')}")
                errors.append(UploadRowError(row_number=index, message="; ".join(issues)))
                continue

            valid_models.append(model)
            valid_rows.append(model.model_dump(by_alias=True))

        return _ValidationResult(
            file_format=file_format,
            valid_models=valid_models,
            valid_rows=valid_rows,
            errors=errors,
            total_rows=len(rows),
        )

    def preview_upload(
        self,
        entity_type: UploadEntityType,
        filename: str,
        content: bytes,
        sample_size: int = 20,
    ) -> UploadPreviewResponse:
        validation = self._validate_rows(entity_type, filename, content)
        return UploadPreviewResponse(
            entity_type=entity_type,
            file_format=validation.file_format,
            total_rows=validation.total_rows,
            valid_rows=len(validation.valid_models),
            invalid_rows=len(validation.errors),
            sample_rows=validation.valid_rows[:sample_size],
            errors=validation.errors[:sample_size],
        )

    async def commit_upload(
        self,
        entity_type: UploadEntityType,
        filename: str,
        content: bytes,
        allow_partial: bool,
    ) -> UploadCommitResponse:
        validation = self._validate_rows(entity_type, filename, content)
        if validation.total_rows == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Upload file contains no data rows.",
            )

        if validation.errors and not allow_partial:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail={
                    "message": "Upload contains invalid rows. Fix errors or use allowPartial=true.",
                    "errors": [error.model_dump() for error in validation.errors[:50]],
                },
            )

        if not validation.valid_models:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="No valid rows found in upload.",
            )

        if entity_type == "portfolios":
            await self._ingestion_service.publish_portfolios(
                [model for model in validation.valid_models if isinstance(model, Portfolio)]
            )
        elif entity_type == "instruments":
            await self._ingestion_service.publish_instruments(
                [model for model in validation.valid_models if isinstance(model, Instrument)]
            )
        elif entity_type == "transactions":
            await self._ingestion_service.publish_transactions(
                [model for model in validation.valid_models if isinstance(model, Transaction)]
            )
        elif entity_type == "market_prices":
            await self._ingestion_service.publish_market_prices(
                [model for model in validation.valid_models if isinstance(model, MarketPrice)]
            )
        elif entity_type == "fx_rates":
            await self._ingestion_service.publish_fx_rates(
                [model for model in validation.valid_models if isinstance(model, FxRate)]
            )
        else:
            await self._ingestion_service.publish_business_dates(
                [model for model in validation.valid_models if isinstance(model, BusinessDate)]
            )

        return UploadCommitResponse(
            entity_type=entity_type,
            file_format=validation.file_format,
            total_rows=validation.total_rows,
            valid_rows=len(validation.valid_models),
            invalid_rows=len(validation.errors),
            published_rows=len(validation.valid_models),
            skipped_rows=len(validation.errors),
            message="Upload committed and queued for processing.",
        )


def get_upload_ingestion_service(
    ingestion_service: IngestionService = Depends(get_ingestion_service),
) -> UploadIngestionService:
    return UploadIngestionService(ingestion_service)
